#!/usr/bin/env python3
"""
Scraper Dispatcher — Polls Google Sheet for new scraping requests and executes them.

Architecture:
  1. Downloads the "Target Scraper Request" Google Sheet via rclone (xlsx export)
  2. Finds rows with Status = "New"
  3. Dispatches to the appropriate scraper:
     - Marketplace → Node.js sequential-ID adapter (or creates one via Claude)
     - Trade Fair  → Claude Code SDK handles recon + scraping
     - Deep Research → Flagged for manual handling
  4. Uploads results to Google Drive
  5. Updates the sheet status

Usage:
  python scrapers/dispatcher.py                 # Process all "New" requests
  python scrapers/dispatcher.py --dry-run       # Show what would be processed
  python scrapers/dispatcher.py --request 7     # Process specific request #

Requirements:
  pip install openpyxl claude-code-sdk
  rclone configured with 'gdrive:' remote
"""

import asyncio
import json
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRAPERS_DIR = PROJECT_ROOT / "scrapers"
MARKETPLACES_DIR = SCRAPERS_DIR / "marketplaces"
OUTPUT_DIR = PROJECT_ROOT / "output"

# Google Sheet location in Google Drive
SHEET_DRIVE_PATH = "GTM Pipeline/Phase 0: Web Crawler Agent/Target Scraper Request"

# Google Drive folder for results upload
RESULTS_DRIVE_FOLDER_ID = "1Fmgyw28S4Xsu5ZRwJBofGXp3BMy1bt8E"

# rclone remote name
RCLONE_REMOTE = "gdrive:"

# Column indices in the sheet (0-based)
COL_REQUEST_NUM = 0
COL_DATE = 1
COL_SUBMITTED_BY = 2
COL_SOURCE_TYPE = 3
COL_EVENT_NAME = 4
COL_URL = 5
COL_COUNTRY = 6
COL_PRIORITY = 7
COL_FIELDS_WANTED = 8
COL_NOTES = 9
COL_STATUS = 10
COL_RESULTS_LINK = 11
COL_COMPANIES_FOUND = 12
COL_COMPLETED_DATE = 13
COL_PROCESSOR_NOTES = 14

# Known marketplace adapters
KNOWN_ADAPTERS = {
    "bq": ["diy.com", "b&q"],
    "mediamarkt": ["mediamarkt"],
}

# Log file
LOG_DIR = PROJECT_ROOT / "output" / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)


def log(msg: str):
    """Print and log a message."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {msg}"
    print(line)
    log_file = LOG_DIR / f"dispatcher_{datetime.now().strftime('%Y%m%d')}.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ── Google Sheet operations ───────────────────────────────────────────────────

def download_sheet() -> Path:
    """Download the Google Sheet as xlsx via rclone."""
    tmp_dir = Path(tempfile.mkdtemp(prefix="scraper_dispatch_"))
    result = subprocess.run(
        ["rclone", "copy", f"{RCLONE_REMOTE}{SHEET_DRIVE_PATH}.xlsx", str(tmp_dir)],
        capture_output=True, text=True, timeout=60,
    )
    if result.returncode != 0:
        raise RuntimeError(f"rclone download failed: {result.stderr}")

    xlsx_files = list(tmp_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise RuntimeError(f"No xlsx file downloaded to {tmp_dir}")

    log(f"Downloaded sheet to {xlsx_files[0]}")
    return xlsx_files[0]


def parse_requests(xlsx_path: Path) -> list[dict]:
    """Parse the Requests sheet and return all rows as dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["Requests"]

    requests = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= COL_STATUS or not row[COL_REQUEST_NUM]:
            continue

        requests.append({
            "row": row_idx,
            "request_num": int(row[COL_REQUEST_NUM]),
            "date": str(row[COL_DATE] or ""),
            "submitted_by": str(row[COL_SUBMITTED_BY] or ""),
            "source_type": str(row[COL_SOURCE_TYPE] or "").strip(),
            "event_name": str(row[COL_EVENT_NAME] or "").strip(),
            "url": str(row[COL_URL] or "").strip(),
            "country": str(row[COL_COUNTRY] or ""),
            "priority": str(row[COL_PRIORITY] or "Normal"),
            "fields_wanted": str(row[COL_FIELDS_WANTED] or ""),
            "notes": str(row[COL_NOTES] or ""),
            "status": str(row[COL_STATUS] or "").strip(),
            "results_link": str(row[COL_RESULTS_LINK] or ""),
            "companies_found": row[COL_COMPANIES_FOUND],
            "completed_date": str(row[COL_COMPLETED_DATE] or ""),
            "processor_notes": str(row[COL_PROCESSOR_NOTES] or ""),
        })

    wb.close()
    return requests


# ── Google Sheets API (direct cell updates via service account) ───────────────

SHEET_ID = "1-fjYFJdx7zVJY6d8WXQpHtGFLJa9go9QmL-iYoy463s"
SA_KEY_PATH = Path("C:/Users/ledoa/Downloads/channelengine-ai-e7e32c629d8d.json")

_gspread_client = None


def get_gspread_client():
    """Lazy-init gspread client with service account."""
    global _gspread_client
    if _gspread_client is None:
        import gspread
        from google.oauth2.service_account import Credentials
        creds = Credentials.from_service_account_file(
            str(SA_KEY_PATH),
            scopes=["https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive"],
        )
        _gspread_client = gspread.authorize(creds)
    return _gspread_client


def update_sheet_status(xlsx_path: Path, row: int, status: str,
                        processor_notes: str = "", companies_found: int = None,
                        results_link: str = ""):
    """Update the status directly in the live Google Sheet via API."""
    try:
        gc = get_gspread_client()
        ws = gc.open_by_key(SHEET_ID).worksheet("Requests")

        ws.update_cell(row, COL_STATUS + 1, status)
        if processor_notes:
            ws.update_cell(row, COL_PROCESSOR_NOTES + 1, processor_notes)
        if companies_found is not None:
            ws.update_cell(row, COL_COMPANIES_FOUND + 1, companies_found)
        if status == "Completed":
            ws.update_cell(row, COL_COMPLETED_DATE + 1,
                           datetime.now().strftime("%Y-%m-%d"))
        if results_link:
            ws.update_cell(row, COL_RESULTS_LINK + 1, results_link)

        log(f"Sheet updated: row {row} -> {status}")
    except Exception as e:
        log(f"WARNING: Failed to update sheet via API: {e}")
        # Fallback: update local xlsx copy
        wb = openpyxl.load_workbook(xlsx_path)
        ws_local = wb["Requests"]
        ws_local.cell(row=row, column=COL_STATUS + 1, value=status)
        if processor_notes:
            ws_local.cell(row=row, column=COL_PROCESSOR_NOTES + 1, value=processor_notes)
        wb.save(xlsx_path)
        wb.close()


def upload_sheet(xlsx_path: Path):
    """No-op: sheet is updated directly via Google Sheets API now."""
    log("Sheet already updated via API (no file upload needed)")


# ── Results upload ────────────────────────────────────────────────────────────

def determine_upload_subfolder(request: dict) -> str:
    """Determine the Drive subfolder based on source type."""
    source = request["source_type"]
    name = request["event_name"].replace("/", "-").replace("\\", "-").strip()

    if source == "Marketplace":
        return f"Marketplace Scraped Data/{name}"
    elif source == "Trade Fair":
        return f"Event Scraped Data/{name}"
    else:
        return f"Other/{name}"


def upload_results(results_path: Path, request: dict) -> str:
    """Upload results to the target Google Drive folder. Returns Drive link or empty string."""
    subfolder = determine_upload_subfolder(request)
    drive_dest = f"{RCLONE_REMOTE}{{{RESULTS_DRIVE_FOLDER_ID}}}/{subfolder}/"

    # Try using --drive-root-folder-id instead
    result = subprocess.run(
        ["rclone", "copy", str(results_path),
         f"{RCLONE_REMOTE}{subfolder}/",
         "--drive-root-folder-id", RESULTS_DRIVE_FOLDER_ID],
        capture_output=True, text=True, timeout=300,
    )
    if result.returncode != 0:
        log(f"WARNING: Drive upload failed: {result.stderr}")
        return ""

    log(f"Results uploaded to Drive: {subfolder}")
    return f"https://drive.google.com/drive/folders/{RESULTS_DRIVE_FOLDER_ID}"


# ── Marketplace scraping ─────────────────────────────────────────────────────

def find_adapter(request: dict) -> str | None:
    """Check if a marketplace adapter already exists for this URL."""
    url = request["url"].lower()
    for adapter_id, patterns in KNOWN_ADAPTERS.items():
        if any(p in url for p in patterns):
            adapter_file = MARKETPLACES_DIR / "marketplaces" / f"{adapter_id}.js"
            if adapter_file.exists():
                return adapter_id
    return None


def run_marketplace_scraper(adapter_id: str, from_id: int = None, to_id: int = None) -> Path:
    """Run a marketplace scraper and return the results directory."""
    cmd = ["node", "scrape.mjs", adapter_id]
    if from_id is not None:
        cmd.extend(["--from", str(from_id)])
    if to_id is not None:
        cmd.extend(["--to", str(to_id)])

    log(f"Running marketplace scraper: {' '.join(cmd)}")
    result = subprocess.run(
        cmd, cwd=str(MARKETPLACES_DIR),
        capture_output=True, text=True, timeout=3600,
    )

    if result.returncode != 0:
        log(f"Scraper error output: {result.stderr[:500]}")
        raise RuntimeError(f"Marketplace scraper failed: {result.stderr[:200]}")

    log(f"Scraper output: {result.stdout[-500:]}")
    results_dir = MARKETPLACES_DIR / "results" / adapter_id
    return results_dir


# ── Claude Code CLI dispatch ──────────────────────────────────────────────────

# Find the Claude Code CLI binary (npm package, NOT the desktop app)
CLAUDE_CLI = None
_candidates = [
    Path(os.environ.get("APPDATA", "")) / "npm" / "claude.cmd",
    Path(os.environ.get("APPDATA", "")) / "npm" / "claude",
    Path(os.environ.get("HOME", "")) / "node_modules" / ".bin" / "claude",
]
for p in _candidates:
    if p.exists():
        CLAUDE_CLI = str(p)
        break


async def dispatch_to_claude(request: dict) -> dict:
    """Use Claude Code CLI (claude -p) to handle a scraping request."""
    if not CLAUDE_CLI:
        return {"status": "failed", "error": "Claude CLI not found. Install with: npm install -g @anthropic-ai/claude-code"}

    source_type = request["source_type"]
    event_name = request["event_name"]
    url = request["url"]
    fields = request["fields_wanted"]
    notes = request["notes"]
    country = request["country"]

    if source_type == "Marketplace":
        prompt = f"""You are a marketplace scraper agent. A new marketplace scraping request has come in.

Marketplace: {event_name}
URL: {url}
Country: {country}
Fields wanted: {fields}
Notes: {notes}

Your task:
1. Analyze the seller page at the URL to understand the data structure
2. Check if a marketplace adapter already exists in scrapers/marketplaces/marketplaces/
3. If not, create a new adapter following the pattern in _template.js (use the /marketplace-with-sequential-id-scrapper skill)
4. Run the scraper with a small test range first (--from 1 --to 50)
5. If the test succeeds, run the full range
6. Save the results CSV path

IMPORTANT: Work from the project root at {PROJECT_ROOT}
The marketplace scraper entry point is: scrapers/marketplaces/scrape.mjs
Adapters go in: scrapers/marketplaces/marketplaces/<name>.js

Output a JSON summary at the end:
{{"status": "completed", "adapter": "<name>", "results_path": "<path>", "companies_found": <n>}}
"""
    elif source_type == "Trade Fair":
        safe_name = event_name.lower().replace(' ', '_').replace('+', '').replace('&', '')
        prompt = f"""DO NOT ASK QUESTIONS. DO NOT PRESENT OPTIONS. JUST DO THE WORK.

Scrape ALL exhibitors from this trade fair and save as CSV. Start immediately.

Event: {event_name}
URL: {url}
Country: {country}
Fields wanted: {fields}
Notes: {notes}

Steps — execute ALL of them now:
1. Use curl or python httpx to fetch the exhibitor listing page at {url}
2. Analyze the HTML/JS to find the data source (look for Algolia, REST APIs, embedded JSON, XHR endpoints)
3. Write a Python script to extract ALL exhibitors with: company name, website, country, contact details
4. Run the script and save results as CSV to output/{safe_name}/exhibitors.csv
5. Print the final line: {{"status": "completed", "results_path": "output/{safe_name}/exhibitors.csv", "companies_found": <COUNT>}}

Work from: {PROJECT_ROOT}
Reference existing scrapers in scrapers/tradefairs/ for patterns.
Do NOT ask for confirmation. Do NOT present options. Execute everything now.
"""
    else:
        return {"status": "skipped", "reason": f"Unsupported source type: {source_type}"}

    log(f"Dispatching to Claude CLI: {source_type} - {event_name}")

    try:
        log(f"Claude CLI path: {CLAUDE_CLI}")

        # Write prompt to temp file to avoid Windows argument length limits
        prompt_file = Path(tempfile.mktemp(suffix=".txt", prefix="claude_prompt_"))
        prompt_file.write_text(prompt, encoding="utf-8")

        # Pipe the prompt via stdin to avoid shell escaping issues
        result = subprocess.run(
            [CLAUDE_CLI, "-p", "-",
             "--allowedTools", "Bash,Read,Write,Edit,Glob,Grep",
             "--max-turns", "50",
             "--output-format", "text",
             "--no-session-persistence"],
            cwd=str(PROJECT_ROOT),
            input=prompt,
            capture_output=True, text=True,
            timeout=1800,  # 30 minute timeout
        )

        prompt_file.unlink(missing_ok=True)

        result_text = result.stdout
        log(f"Claude CLI exit code: {result.returncode}")
        log(f"Claude CLI stdout (last 500 chars): {result_text[-500:]}")
        if result.stderr:
            log(f"Claude CLI stderr (last 500 chars): {result.stderr[-500:]}")

        if result.returncode != 0 and not result_text:
            return {"status": "failed", "error": result.stderr[:200]}

    except subprocess.TimeoutExpired:
        return {"status": "failed", "error": "Claude CLI timed out after 30 minutes"}
    except Exception as e:
        log(f"Claude CLI error: {e}")
        return {"status": "failed", "error": str(e)}

    # Try to parse JSON summary from the result
    try:
        import re
        json_matches = re.findall(r'\{[^{}]*"status"[^{}]*\}', result_text)
        if json_matches:
            return json.loads(json_matches[-1])
    except (json.JSONDecodeError, IndexError):
        pass

    return {"status": "completed", "raw_output": result_text[-500:]}


# ── Main dispatcher ──────────────────────────────────────────────────────────

async def process_request(request: dict, xlsx_path: Path) -> dict:
    """Process a single scraping request."""
    req_num = request["request_num"]
    source_type = request["source_type"]
    event_name = request["event_name"]

    log(f"Processing request #{req_num}: {source_type} - {event_name}")

    # Update status to "In Progress"
    update_sheet_status(xlsx_path, request["row"], "In Progress",
                        processor_notes=f"Auto-dispatcher started at {datetime.now().strftime('%H:%M')}")

    try:
        # Check if this is a known marketplace with an existing adapter
        if source_type == "Marketplace":
            adapter_id = find_adapter(request)
            if adapter_id:
                log(f"Found existing adapter: {adapter_id}")
                results_dir = run_marketplace_scraper(adapter_id)
                csv_file = results_dir / "sellers.csv"
                if csv_file.exists():
                    import csv as csv_mod
                    with open(csv_file, "r", encoding="utf-8") as f:
                        count = sum(1 for _ in csv_mod.reader(f)) - 1  # minus header

                    # Upload results
                    drive_link = upload_results(results_dir, request)

                    update_sheet_status(
                        xlsx_path, request["row"], "Completed",
                        processor_notes=f"Auto-scraped with {adapter_id} adapter. {count} sellers found.",
                        companies_found=count,
                        results_link=drive_link,
                    )
                    return {"status": "completed", "adapter": adapter_id, "companies_found": count}

        # Dispatch to Claude Code SDK for unknown marketplaces and trade fairs
        if source_type in ("Marketplace", "Trade Fair"):
            result = await dispatch_to_claude(request)

            if result.get("status") == "completed":
                results_path = result.get("results_path", "")
                companies = result.get("companies_found", 0)
                drive_link = ""

                if results_path and Path(results_path).exists():
                    drive_link = upload_results(Path(results_path), request)

                update_sheet_status(
                    xlsx_path, request["row"], "Completed",
                    processor_notes=f"Auto-scraped via Claude. {companies} companies found.",
                    companies_found=companies,
                    results_link=drive_link,
                )
            else:
                error = result.get("error", result.get("reason", "Unknown error"))
                update_sheet_status(
                    xlsx_path, request["row"], "Failed",
                    processor_notes=f"Auto-dispatcher error: {error[:200]}")

            return result

        # Deep Research — skip for now
        if source_type == "Deep Research":
            log(f"Skipping Deep Research request #{req_num} (requires manual handling)")
            update_sheet_status(
                xlsx_path, request["row"], "Needs Review",
                processor_notes="Deep Research requests require manual handling.")
            return {"status": "skipped", "reason": "Deep Research not automated"}

        # Unknown source type
        log(f"Unknown source type '{source_type}' for request #{req_num}")
        return {"status": "skipped", "reason": f"Unknown source type: {source_type}"}

    except Exception as e:
        log(f"Error processing request #{req_num}: {e}")
        update_sheet_status(
            xlsx_path, request["row"], "Failed",
            processor_notes=f"Auto-dispatcher error: {str(e)[:200]}")
        return {"status": "failed", "error": str(e)}


async def main():
    """Main entry point."""
    import argparse
    parser = argparse.ArgumentParser(description="Scraper Dispatcher")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be processed")
    parser.add_argument("--request", type=int, help="Process specific request number")
    args = parser.parse_args()

    log("=" * 60)
    log("Scraper Dispatcher starting")
    log("=" * 60)

    # Step 1: Download the sheet
    try:
        xlsx_path = download_sheet()
    except Exception as e:
        log(f"FATAL: Could not download sheet: {e}")
        sys.exit(1)

    # Step 2: Parse requests
    requests = parse_requests(xlsx_path)
    log(f"Found {len(requests)} total requests in sheet")

    # Step 3: Filter to actionable requests
    if args.request:
        actionable = [r for r in requests if r["request_num"] == args.request]
        if not actionable:
            log(f"Request #{args.request} not found")
            sys.exit(1)
    else:
        actionable = [r for r in requests if r["status"] == "New"]

    if not actionable:
        log("No new requests to process. Done.")
        return

    log(f"\n{'='*60}")
    log(f"Found {len(actionable)} request(s) to process:")
    for r in actionable:
        log(f"  #{r['request_num']} | {r['source_type']:<15} | {r['event_name']:<25} | {r['priority']}")
    log(f"{'='*60}\n")

    if args.dry_run:
        log("DRY RUN — no scraping will be performed.")
        return

    # Step 4: Process each request (Urgent first)
    priority_order = {"Urgent": 0, "Normal": 1}
    actionable.sort(key=lambda r: priority_order.get(r["priority"], 2))

    results_summary = []
    for request in actionable:
        result = await process_request(request, xlsx_path)
        results_summary.append({
            "request": request["request_num"],
            "event": request["event_name"],
            "result": result.get("status", "unknown"),
        })

    # Step 5: Upload updated sheet back to Drive
    try:
        upload_sheet(xlsx_path)
    except Exception as e:
        log(f"WARNING: Could not upload updated sheet: {e}")

    # Summary
    log(f"\n{'='*60}")
    log("DISPATCH SUMMARY")
    log(f"{'='*60}")
    for s in results_summary:
        log(f"  #{s['request']} {s['event']:<25} -> {s['result']}")
    log(f"{'='*60}")
    log("Dispatcher finished.")


if __name__ == "__main__":
    asyncio.run(main())
